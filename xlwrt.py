#!/usr/bin/env python
# -*- coding: utf-8 -*-

def extract_col(cell_idx):
    import re
    return re.match(r"[A-Za-z]+", cell_idx).group()
def extract_row(cell_idx):
    import re
    return re.match(r".*(\d+).*", 'A1').group(1)

def to_col_num(col_alpha):
    i =  0
    col = 0
    for l in [(ord(a)-ord('A') + 1) for a in col_alpha[::-1]]:
        col += l * pow(26,i)
        i+=1
    return col

def parse_args():
    import argparse
    parser = argparse.ArgumentParser(description='Writes csv read from standard input to a excel file.')
    parser.add_argument('--top_left_cell','-c', nargs=1, metavar='A1')
    parser.add_argument('--xlsx','-x', nargs=1, metavar='excel_file.xlsx')
    parser.add_argument('--sheet','-s', nargs=1, metavar='sheet_name')
    return parser.parse_args()

def get_sheet(xlsx, sheet_name):
    # http://openpyxl.readthedocs.org/en/default/tutorial.html
    return load_workbook(xlsx).get_sheet_by_name(sheet_name)

if __name__ == '__main__':
    import sys
    import csv
    lines = [ l for l in csv.reader(sys.stdin.readlines())]
    row_siz=len(lines)
    col_siz=len(lines[0])

    args = parse_args()

    from openpyxl import load_workbook
    wb = load_workbook(args.xlsx[0])
    sheet = wb.get_sheet_by_name(args.sheet[0])

    l_row = int(extract_row(args.top_left_cell[0]))
    l_col = int(to_col_num(extract_col(args.top_left_cell[0])))
    for (cell_r,txt_r, cell_c,txt_c) in [(r, r - l_row, c, c - l_col)
                         for r in range(l_row, l_row+row_siz)
                         for c in range(l_col, l_col+col_siz)]:
        val = lines[txt_r][txt_c]
        sheet.cell(row=cell_r, column=cell_c).value=int(val) if val.isdigit() else val
    wb.save(args.xlsx[0])
